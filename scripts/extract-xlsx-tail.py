#!/usr/bin/env python3
"""
Extract the XLSX Citations rows that didn't make it into the imported
quotes.json. These typically have one of:

  · A UUID or work-name in the Sujet column instead of a real topic
    label, so the topic resolver returned nothing and the row was
    skipped during import.
  · No Sujet at all, just author + text fragments.
  · A non-numeric value in the ID column (rows 799-802 are placeholder
    citations · the ID column carries the citation reference itself
    rather than an id).

Writes `quotes-from-xlsx-tail.csv` with the same column layout as
`quotes-from-epub.csv` so the two can be edited side-by-side.

Run:
  python3 scripts/extract-xlsx-tail.py
"""
import json
import csv
import openpyxl
import re
import unicodedata
from pathlib import Path

XLSX = '/Users/Janvier/Library/Mobile Documents/com~apple~CloudDocs/for-the-kingdom/fathers/excel/fathers_db.xlsx'
AUTHORS_PATH = Path('src/lib/data/authors.json')
QUOTES_PATH = Path('src/lib/data/quotes.json')
OUT_CSV = Path('quotes-from-xlsx-tail.csv')


def norm_name(s):
    s = unicodedata.normalize('NFD', s or '')
    s = ''.join(c for c in s if not unicodedata.combining(c)).lower()
    s = re.sub(r"[‘’ʼʻ`']", "'", s)
    s = re.sub(r'^(pape\s+)?(st\.?|saint|sainte|ste\.?)\s+', '', s, flags=re.I)
    s = re.sub(r'^pape\s+', '', s, flags=re.I)
    s = re.sub(r"[^a-z0-9']+", ' ', s).strip()
    return s


def main():
    authors = json.loads(AUTHORS_PATH.read_text())
    quotes = json.loads(QUOTES_PATH.read_text())
    imported_ids = {q['id'] for q in quotes}

    author_by_key = {}
    for a in authors:
        for n in [a['name'], a.get('originalName') or '']:
            k = norm_name(n)
            if k:
                author_by_key.setdefault(k, a)

    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb['Citations']

    rows = []
    for r in range(3, ws.max_row + 1):
        row = [c.value for c in ws[r]]
        if not any(v is not None and str(v).strip() for v in row):
            continue
        qid_raw = row[2]
        qid = None
        if isinstance(qid_raw, (int, float)):
            try:
                qid = int(qid_raw)
            except Exception:
                pass
        if qid is not None and qid in imported_ids:
            continue

        # Map XLSX columns to CSV fields (cf. column header dump):
        #   1 OK?  2 ID  3 Auteur  4 Source  5 Sujet
        #   6 EN   7 LA  8 GR/SY   9 Context 10 Link  11 Archive
        #  12 Migne 13 Ref 14 FR   15 EN-translated  16 Notes  17 Cit-Abbr
        auteur = str(row[3]).strip() if row[3] else ''
        source = str(row[4]).strip() if row[4] else ''
        topic = str(row[5]).strip() if row[5] else ''
        en = str(row[6]).strip() if row[6] else ''
        latin = str(row[7]).strip() if row[7] else ''
        greek = str(row[8]).strip() if row[8] else ''
        migne = str(row[12]).strip() if row[12] else ''
        ref = str(row[13]).strip() if row[13] else ''
        fr = str(row[14]).strip() if row[14] else ''
        notes = str(row[16]).strip() if row[16] else ''

        # If the ID cell carries free text (rows 799-802 use it for the
        # citation reference itself), surface that in notes.
        if qid_raw is not None and qid is None:
            notes = (notes + ' · ' if notes else '') + f'(citation from ID column: {qid_raw!r})'

        # Detect the reason the import dropped this row · annotate.
        reasons = []
        if not topic:
            reasons.append('no topic')
        elif re.match(r'^[0-9a-f]{8}-[0-9a-f]{4}-', topic, re.I):
            reasons.append('topic cell holds a UUID, not a label')
        if not auteur:
            reasons.append('no author')
        if reasons:
            notes = (notes + ' · ' if notes else '') + 'skipped: ' + ', '.join(reasons)

        # Try author match
        author_match = ''
        if auteur:
            m = author_by_key.get(norm_name(auteur))
            if m:
                author_match = f"{m['id']} · {m['name']}"

        rows.append({
            'existing': '',
            'existing_id': '',
            'topic_en': topic,  # might already be French or a UUID · user will edit
            'source_kind': 'author' if auteur else ('work' if source else ''),
            'source': auteur or source,
            'author_match': author_match,
            'work': source if auteur else '',
            'reference': ref,
            'date': '',
            'en': en,
            'fr': fr,
            'latin': latin,
            'greek': greek,
            'migne': migne,
            'notes': notes,
        })

    headers = [
        'existing', 'existing_id',
        'topic_en', 'source_kind', 'source', 'author_match',
        'work', 'reference', 'date',
        'en', 'fr', 'latin', 'greek', 'migne', 'notes',
    ]
    with OUT_CSV.open('w', newline='', encoding='utf-8-sig') as f:
        w = csv.DictWriter(f, fieldnames=headers)
        w.writeheader()
        for r in rows:
            w.writerow(r)
    print(f'wrote {OUT_CSV} · {len(rows)} unimported XLSX rows')


if __name__ == '__main__':
    main()
